# checklist_lib.py
import time, datetime as dt
import pandas as pd
from openpyxl import load_workbook
from google.api_core.exceptions import ResourceExhausted
import src.logger as log
import os

lg = log.get("checklist")

def fill_template(template_path: str, qa_chain, out_dir="output") -> str:
    """
    Read Excel file with 'Question' and 'Answer' columns.
    Use QA chain to populate answers and citations. Save to new file with timestamp.
    Returns: path to filled Excel file.
    """
    lg.info(f"Opening template → {template_path}")
    wb = load_workbook(template_path)
    ws = wb.active

    # Detect columns
    header = {c.value.strip(): idx + 1 for idx, c in enumerate(ws[1])}
    tq_col = header.get("Question")
    ans_col = header.get("Answer")
    src_col = header.get("Source")

    if not (tq_col and ans_col):
        raise ValueError("Headers 'Question' and 'Answer' not found")

    if not src_col:
        # If "Source" column doesn't exist, create it at the end
        src_col = ws.max_column + 1
        ws.cell(row=1, column=src_col).value = "Source"
        lg.info("No 'Source' column found — created one.")

    for row in range(2, ws.max_row + 1):
        q = ws.cell(row=row, column=tq_col).value
        if not q:
            continue
        lg.info(f"→ {q}")
        wait = 6
        while True:
            try:
                result = qa_chain({"query": q})
                break
            except ResourceExhausted:
                lg.warning("Rate-limit; sleeping…")
                time.sleep(wait)
                wait = min(wait * 2, 60)

        # Fill Answer
        ans = result.get("result", "")
        ws.cell(row=row, column=ans_col).value = ans
        lg.info(f"✓ {ans[:50]}")

        # Fill Source (pages, or N/A)
        sources = result.get("source_documents", [])
        if sources:
            pages = set()
            for doc in sources:
                meta = doc.metadata
                page = meta.get("page", None)
                if page is not None:
                    pages.add(str(page))
            source_str = ", ".join(sorted(pages)) if pages else "Yes"
        else:
            source_str = "N/A"
        ws.cell(row=row, column=src_col).value = source_str

    # Create new filename with timestamp
    os.makedirs(out_dir, exist_ok=True)
    base_name = os.path.splitext(os.path.basename(template_path))[0]
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"{base_name}_filled_{ts}.xlsx")

    wb.save(out_path)
    lg.info(f"✅ Filled file saved → {out_path}")
    return out_path
